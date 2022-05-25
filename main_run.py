# Kenny Schlegel, Peer Neubert, Peter Protzel
#
# HDC-MiniROCKET : Explicit Time Encoding in Time Series Classification with Hyperdimensional Computing
# Copyright (C) 2022 Chair of Automation Technology / TU Chemnitz

from data.dataset_utils import *
import logging
from models.HDC_MINIROCKET import HDC_MINIROCKET_model
from models.MINIROCKET import MINIROCKET_model

# self.config logger
logger = logging.getLogger('log')

class NetTrial():
    '''
    high level class for training and evaluating 
    '''
    
    def __init__(self,args):           
        # set self.config parameter
        exec('self.config = ' + args.config + '()')
        self.config.HDC_dim = args.HDC_dim
        self.config.scale = args.scale
        self.config.dataset = args.dataset
        self.config.model_name = args.model
        self.config.stat_iterations = args.stat_iterations
        self.config.ucr_idx = args.ucr_idx
        self.config.normalize_data = args.normalize

        # load data at initialization
        self.load_data()

    def load_data(self):
        # load dataset initially
        self.data = load_dataset(self.config.dataset, self.config)
        X_train = self.data[0]
        X_test = self.data[1]
        y_train = self.data[2]
        y_test = self.data[3]

        self.config.n_classes = len(np.unique(y_train))
        self.config.n_inputs = X_train.shape[1]
        self.config.n_steps = X_train.shape[2]

        # if train test data not a list, create one (k-fold data set loading returns a list of splits - therefore we
        # are handling all training set as lists, even if they only contain one set)
        if type(X_train) == list:
            print("given data is not a list")
            self.X_train_list = X_train
            self.X_test_list = X_test
            self.y_train_list = y_train
            self.y_test_list = y_test
        else:
            self.X_train_list = [X_train]
            self.X_test_list = [X_test]
            self.y_train_list = [y_train]
            self.y_test_list = [y_test]

        if self.config.model_name == "HDC_MINIROCKET":
            self.model = HDC_MINIROCKET_model(self.config)
        elif self.config.model_name == "MINIROCKET":
            self.model = MINIROCKET_model(self.config)

    def train(self):
        '''
        training procedure
        @return:
        '''

        acc_stat = []
        f1_stat = []

        # statistical iteration
        for stat_it in range(self.config.stat_iterations):
            logger.info('Statistial iteration: ' + str(stat_it))
    
            # train for each element in list (that is why we need list form, even if it contains only one element)
            logger.info('Training data contains ' + str(len(self.X_train_list)) + ' training instances...')
            f1s = []
            accs = []
    
            for it in range(len(self.X_train_list)):
                try:
                    logger.info(('.......'))
                    logger.info('instance ' + str(it) + ':')

                    X_train = self.X_train_list[it]
                    X_test = self.X_test_list[it]
                    y_train = self.y_train_list[it]
                    y_test = self.y_test_list[it]

                    # self.config.HDC_dim = X_train.shape[1]
                    logger.info('Training dataset shape: ' + str(X_train.shape) + str(y_train.shape))
                    logger.info('Test dataset shape: ' + str(X_test.shape) + str(y_test.shape))

                    self.config.train_count = len(X_train)
                    self.config.test_data_count = len(X_test)
                    self.model.config = self.config

                    # train the model
                    self.model.train_model(X_train,y_train,X_test,y_test)

                    # evaluate the model
                    acc, f1, confusion_matrix = self.model.eval_model(X_train,y_train,X_test,y_test)

                    accs.append(acc)
                    f1s.append(f1)
                except Exception as e:
                    logger.info(e)

            acc_stat.append(np.mean(accs))
            f1_stat.append((np.mean(f1s)))

            logger.info('Accuracy results of statistical repetitions: ' + str(acc_stat))
            logger.info('F1 scores of statistical repetitions: ' + str(f1_stat))

            # write all scores to extra file
            logger.info('Mean Score: ' + str(np.mean(f1_stat)))
            logger.info('Mean Accuracy: ' + str(np.mean(acc_stat)))

            try:
                with open('results/results_' + self.config.dataset + '_' + self.config.model_name + '.txt', 'a') as file:
                    file.write(str(self.config.ucr_idx) + '\t'
                               + str(round(np.mean(f1_stat), 4)) + '\t'
                               + str(round(np.mean(acc_stat), 4)) + '\t'
                               + str(X_train.shape[0]) + '\t'
                               + str(X_train.shape[1]) + '\t'
                               + str(X_train.shape[2]) + '\t'
                               + str(self.config.n_classes) + '\t'
                               + str(round(np.std(f1_stat), 2)) + '\t'
                               + str(round(np.std(acc_stat), 2)) + '\t'
                               + str(self.config.scale) + '\n'
                               )

                with open('results/computing_time_training_' + self.config.dataset + '_' + self.config.model_name + '.txt', 'a') as file:
                    file.write(str(self.config.ucr_idx) + '\t' +
                               str(self.model.train_preproc) + '\t'
                               + str(self.model.training_time) + '\n'
                               )
                with open('results/computing_time_inference_' + self.config.dataset + '_' + self.config.model_name + '.txt',
                          'a') as file:
                    file.write(str(self.config.ucr_idx) + '\t' +
                               str(self.model.test_preproc) + '\t'
                               + str(self.model.testing_time) + '\n'
                               )

                # write results to excel
                file = 'results/' + self.config.model_name + '_results.xlsx'
                acc_df = pd.DataFrame({'data': acc}, index=[0])
                ucr_idx_df = pd.DataFrame({'data': self.config.ucr_idx}, index=[0])
                # write index
                append_df_to_excel(file, pd.DataFrame({'UCR_idx': []}),
                                   sheet_name=self.config.dataset,
                                   startcol=0, index=False, header=True,
                                   startrow=0)
                # write data
                if self.config.best_scale:
                    header_name = 'best_scale_x_val'
                    append_df_to_excel(file, pd.DataFrame({header_name: []}),
                                       sheet_name=self.config.dataset,
                                       startcol=self.config.scales.shape[0]+2, index=False, header=True,
                                       startrow=0)
                    append_df_to_excel(file, acc_df, sheet_name=self.config.dataset,
                                       index=False, header=False, startrow=self.config.ucr_idx + 1,
                                       startcol=self.config.scales.shape[0] + 2)
                elif not self.config.best_scale:
                    header_name = 'scale_idx' + str(self.config.scale_idx)
                    append_df_to_excel(file, pd.DataFrame({header_name: []}),
                                       sheet_name=self.config.dataset,
                                       startcol=self.config.scale_idx + 1, index=False, header=True,
                                       startrow=0)
                    append_df_to_excel(file, acc_df, sheet_name=self.config.dataset,
                                       index=False, header=False, startrow=self.config.ucr_idx + 1,
                                       startcol=self.config.scale_idx+1)
                append_df_to_excel(file, ucr_idx_df, sheet_name=self.config.dataset,
                                   startcol=0, index=False, header=False,
                                   startrow=self.config.ucr_idx + 1)

                # write time measure results to excel
                file_time = 'results/' + self.config.model_name + '_time_results.xlsx'
                time_df = pd.DataFrame({'train_time_preproc': self.model.train_preproc, 'train_time':self.model.training_time,
                                        'test_time_preproc':self.model.test_preproc,'test_time':self.model.testing_time}, index=[0])
                ucr_idx_df = pd.DataFrame({'data': self.config.ucr_idx}, index=[0])
                # write index
                append_df_to_excel(file_time, pd.DataFrame({'UCR_idx': []}),
                                   sheet_name=self.config.dataset,
                                   startcol=0, index=False, header=True,
                                   startrow=0)
                # write data
                append_df_to_excel(file_time, time_df, sheet_name=self.config.dataset,
                                   index=False, header=False, startrow=self.config.ucr_idx + 1,
                                   startcol=1)
                append_df_to_excel(file_time, time_df.drop(time_df.index), sheet_name=self.config.dataset,
                                   index=False, header=True, startrow=0,
                                   startcol=1)
                append_df_to_excel(file_time, ucr_idx_df, sheet_name=self.config.dataset,
                                   startcol=0, index=False, header=False,
                                   startrow=self.config.ucr_idx + 1)


            except Exception as e:
                logger.info(e)

    def eval(self):
        '''
        evaluating procedure
        @return:
        '''

        f1s = []
        accs = []
        for it in range(len(self.X_test_list)):
            logger.info(('.......'))
            logger.info('instance ' + str(it) + ':')

            X_train = self.X_train_list[it]
            X_test = self.X_test_list[it]
            y_train = self.y_train_list[it]
            y_test = self.y_test_list[it]

            logger.info('Test dataset shape: ' + str(X_test.shape) + str(y_test.shape))

            self.config.test_data_count = len(X_test)
            self.model.config = self.config

            # evaluate the model
            acc, f1, confusion_matrix = self.model.eval_model(X_train,y_train,X_test,y_test,fold=it)

            accs.append(acc)
            f1s.append(f1)

        with open('results/computing_time_inference_' + self.config.dataset + '_' + self.config.model_name + '.txt',
                  'a') as file:
            file.write(str(self.config.ucr_idx) + '\t' +
                       str(self.model.test_preproc) + '\t'
                       + str(self.model.testing_time) + '\n'
                       )

        logger.info('Accuracy results: ' + str(np.mean(accs)))
        logger.info('F1 scores: ' + str(np.mean(f1s)))



def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    code from https://gist.github.com/fredpedroso/590e54d4f07d0ae2d20d0ec0b190d5ff

    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()